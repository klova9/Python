from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

data = Reference(sheet,
          min_col=min_column+1,
          max_col=max_column,
          min_row=min_row,
          max_row=max_row)
catagories = Reference(sheet,
          min_col=1,
          max_col=1,
          min_row=6,
          max_row=7)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(catagories)
sheet.add_chart(barchart, 'B12')
barchart.title = 'Sales by Product line'
barchart.style = 10
wb.save('bar_chart.xlsx')
