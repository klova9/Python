from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = load_workbook('report.xlsx')
sheet = wb['Report']
dim_holder = DimensionHolder(worksheet=sheet)
month = 'January'

sheet['A1'] = 'Sales Report'
sheet['A2'] = 'January'
sheet['A1'].font = Font('Arial', size=20, bold=True)
sheet['A2'].font = Font('Arial', size=14, bold=True)
for col in range(sheet.min_column, sheet.max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

sheet.column_dimensions = dim_holder
wb.save(f'{month}_report.xlsx')